import re
import openpyxl
import csv

# 打开Excel文件
workbook = openpyxl.load_workbook("待填入表格.xlsx")

# 选择工作表
worksheet = workbook.active

# 定义关键字及其对应的正则表达式
keywords = {
    "技术员姓名": r"技术员姓名|技术\s*员\s*姓名",
    "质检员姓名": r"质检员姓名|质检\s*员\s*姓名",
    "安全员姓名": r"安全员姓名|安全\s*员\s*姓名",
    "环保员姓名": r"环保员姓名|环保\s*员\s*姓名",
    "监理员姓名": r"监理员姓名|监理\s*员\s*姓名",
    "施工开始时间": r"施工\s*开始\s*时间|Construction\s*Start\s*Time",
    "施工完成时间": r"施工\s*完成\s*时间|Construction\s*Completion\s*Time",
    "混泥土施工配料单编号": r"混泥土\s*施工\s*配料单\s*编号|Concrete\s*Construction\s*Batch\s*Number",
    "材料种类1": r"材料\s*种类\s*1|Material\s*Type\s*1",
    "生产厂家1": r"生产\s*厂家\s*1|Manufacturer\s*1",
    "材料种类2": r"材料\s*种类\s*2|Material\s*Type\s*2",
    "生产厂家2": r"生产\s*厂家\s*2|Manufacturer\s*2",
    "是否发生变更": r"是否\s*发生\s*变更|Changed\s*",
}
# 创建一个空的数据列表
data_list = []
with open("待导入数据.csv", "r") as csv_file:
    csv_reader = csv.reader(csv_file)
    for row in csv_reader:
        if len(row) == 2:
            text = f"{row[0].strip()}: {row[1].strip()}"
            # 将提取的数据按行拼接成text格式
            lines = text.strip().split("\n")
            data = {}  # 创建一个空的数据字典
            for line in lines:
                parts = line.split(":")
                if len(parts) == 2:
                    key = parts[0].strip()
                    value = parts[1].strip()
                    for keyword, pattern in keywords.items():
                        if re.search(pattern, key, re.IGNORECASE):
                            data[keyword] = value
            data_list.append(data)  # 将每组数据添加到数据列表中

# 将提取的数据填充到Excel表格
for data in data_list:
    for keyword, value in data.items():
        # 查找关键字所在的列
        a=0
        for cell in worksheet[1]:
            if cell.value == keyword:
                a=1
                column = cell.column
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row+1, min_col=column, max_col=column):
                    #row=(2,3,4,5)
                    for cell in row:
                        b=0
                        if  cell.value is None:
                            cell.value = value
                            b=1
                            break  # 找到空格后跳出内层循环
                    if b == 1:
                        break
        if a == 0:
            print(f"未找到 {keyword}：{value} 所在的列")

# 保存Excel文件
workbook.save("导入完成表格.xlsx")
